"""
Jira User Story checklist automation.

Reads an Excel file listing user stories (one sheet per team, or a Team column),
fetches each issue from Jira, runs the checklist verifications from
"TQ checklist Jira US ready check", optionally auto-updates fields or emails
the team leader, and writes a report with PASS/FAIL, failing steps, execution
time and runtime errors.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import smtplib
import sys
import time
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
)
log = logging.getLogger("jira_checker")


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class StepResult:
    name: str
    passed: bool
    message: str = ""
    action: str = ""  # "updated", "emailed", "skipped", "none"
    error: str = ""
    expected: str = ""
    actual: str = ""


@dataclass
class IssueResult:
    issue_key: str
    team: str
    overall: str = "PASS"                 # PASS or FAIL
    duration_seconds: float = 0.0
    runtime_error: str = ""
    steps: list[StepResult] = field(default_factory=list)

    def failing_step_names(self) -> list[str]:
        return [s.name for s in self.steps if not s.passed]


# ---------------------------------------------------------------------------
# Configuration loading
# ---------------------------------------------------------------------------

def load_config(path: Path) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Jira client
# ---------------------------------------------------------------------------

class JiraClient:
    def __init__(self, base_url: str, bearer_token: str, verify_ssl: bool = True, timeout: int = 30):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update({
            "Authorization": f"Bearer {bearer_token}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        })
        self.session.verify = verify_ssl
        self.timeout = timeout

    def get_issue(self, key: str) -> dict[str, Any]:
        url = f"{self.base_url}/rest/api/2/issue/{key}"
        r = self.session.get(url, timeout=self.timeout)
        r.raise_for_status()
        return r.json()

    def update_issue(self, key: str, fields: dict[str, Any]) -> None:
        url = f"{self.base_url}/rest/api/2/issue/{key}"
        payload = {"fields": fields}
        r = self.session.put(url, json=payload, timeout=self.timeout)
        if r.status_code >= 400:
            raise RuntimeError(f"Jira update failed ({r.status_code}): {r.text}")


# ---------------------------------------------------------------------------
# Emailer
# ---------------------------------------------------------------------------

class Mailer:
    def __init__(self, smtp_cfg: dict[str, Any], dry_run: bool = False):
        self.cfg = smtp_cfg
        self.dry_run = dry_run

    def send(self, to_address: str, subject: str, body: str,
             html_body: str | None = None) -> None:
        if self.dry_run or not to_address:
            log.info("[DRY-RUN email] to=%s subject=%s html=%s",
                     to_address, subject, bool(html_body))
            return
        msg = EmailMessage()
        msg["From"] = self.cfg["from_address"]
        msg["To"] = to_address
        msg["Subject"] = subject
        msg.set_content(body)
        if html_body:
            msg.add_alternative(html_body, subtype="html")

        with smtplib.SMTP(self.cfg["host"], self.cfg["port"]) as s:
            if self.cfg.get("use_tls"):
                s.starttls()
            user = self.cfg.get("username") or os.getenv("SMTP_USERNAME")
            pwd = self.cfg.get("password") or os.getenv("SMTP_PASSWORD")
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)


# ---------------------------------------------------------------------------
# Individual checks
# ---------------------------------------------------------------------------

def _has_value(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    if isinstance(v, (list, dict)):
        return len(v) > 0
    return True


def _list_values(items: list[dict] | None, key: str = "name") -> list[str]:
    if not items:
        return []
    return [i.get(key, "") for i in items]


def _option_value(v: Any) -> str:
    """Render a Jira field value (option dict, list of options, scalar) for the report."""
    if v is None:
        return "<empty>"
    if isinstance(v, list):
        if not v:
            return "<empty>"
        return ", ".join(_option_value(x) for x in v)
    if isinstance(v, dict):
        return str(v.get("value") or v.get("name") or v)
    return str(v)


def check_contains_named(item_list: list[dict] | None, expected_name: str) -> bool:
    return expected_name in _list_values(item_list or [], "name")


class IssueProcessor:
    def __init__(self, jira: JiraClient, mailer: Mailer, config: dict[str, Any],
                 auto_update: bool = True, send_emails: bool = True,
                 raw_response_dir: Path | None = None):
        self.jira = jira
        self.mailer = mailer
        self.cfg = config
        self.auto_update = auto_update
        self.send_emails = send_emails
        self.field_ids = config["field_ids"]
        self.expected = config["expected_values"]
        self.teams = config["teams"]
        self.raw_response_dir = raw_response_dir
        if raw_response_dir is not None:
            raw_response_dir.mkdir(parents=True, exist_ok=True)

    # --- helpers -----------------------------------------------------------

    def _update(self, key: str, fields: dict[str, Any], result: StepResult) -> None:
        if not self.auto_update:
            result.action = "update-skipped (disabled)"
            return
        try:
            self.jira.update_issue(key, fields)
            result.action = f"updated:{list(fields.keys())}"
        except Exception as e:  # noqa: BLE001
            result.action = "update-failed"
            result.error = str(e)

    def _check_single_option_field(self, key: str, fields: dict, step_name: str,
                                   field_id: str, expected: str) -> StepResult:
        """Strict-equality check for single-value option fields.

        empty    -> FAIL + auto-fill with expected
        equal    -> PASS
        mismatch -> FAIL (no overwrite — value may be intentional, requires review)
        """
        v = fields.get(field_id)
        actual = _option_value(v)
        if not _has_value(v):
            res = StepResult(step_name, False, "Empty",
                             expected=expected, actual=actual)
            self._update(key, {field_id: [{"value": expected}]}, res)
            return res
        if actual == expected:
            return StepResult(step_name, True, "Matches expected",
                              expected=expected, actual=actual)
        return StepResult(
            step_name, False,
            f"Value mismatch: expected '{expected}', got '{actual}'",
            expected=expected, actual=actual,
        )

    # --- check methods ------------------------------------------------------

    def check_affects_version(self, key: str, fields: dict) -> StepResult:
        expected = self.expected["affects_version"]
        current = _list_values(fields.get("versions"), "name")
        actual = ", ".join(current) if current else "<empty>"
        if expected in current:
            return StepResult("Affects version", True, f"Contains '{expected}'",
                              expected=expected, actual=actual)
        res = StepResult("Affects version", False, f"Missing '{expected}'",
                         expected=expected, actual=actual)
        new_versions = [{"name": v} for v in current] + [{"name": expected}]
        self._update(key, {"versions": new_versions}, res)
        return res

    def check_fix_version(self, key: str, fields: dict) -> StepResult:
        expected = self.expected["fix_version"]
        current = _list_values(fields.get("fixVersions"), "name")
        actual = ", ".join(current) if current else "<empty>"
        if expected in current:
            return StepResult("Fix version", True, f"Contains '{expected}'",
                              expected=expected, actual=actual)
        res = StepResult("Fix version", False, f"Missing '{expected}'",
                         expected=expected, actual=actual)
        new_versions = [{"name": v} for v in current] + [{"name": expected}]
        self._update(key, {"fixVersions": new_versions}, res)
        return res

    def check_component(self, key: str, fields: dict) -> StepResult:
        expected = self.expected["component"]
        current = _list_values(fields.get("components"), "name")
        actual = ", ".join(current) if current else "<empty>"
        if expected in current:
            return StepResult("Component", True, f"Contains '{expected}'",
                              expected=expected, actual=actual)
        res = StepResult("Component", False, f"Missing '{expected}'",
                        expected=expected, actual=actual)
        new_components = [{"name": c} for c in current] + [{"name": expected}]
        self._update(key, {"components": new_components}, res)
        return res

    def check_labels(self, key: str, fields: dict, team: str) -> StepResult:
        base_label = self.expected["base_label"]
        current_labels = fields.get("labels") or []
        missing: list[str] = []
        if base_label not in current_labels:
            missing.append(base_label)

        team_info = self.teams.get(team)
        team_label = team_info.get("label") if team_info else None
        if team_label and team_label not in current_labels:
            missing.append(team_label)

        expected_str = base_label + (f", {team_label}" if team_label else "")
        actual = ", ".join(current_labels) if current_labels else "<empty>"

        if not missing:
            return StepResult("Labels", True,
                              f"Has {base_label}" + (f" + {team_label}" if team_label else ""),
                              expected=expected_str, actual=actual)

        res = StepResult("Labels", False, f"Missing labels: {missing}",
                         expected=expected_str, actual=actual)
        new_labels = current_labels + missing
        self._update(key, {"labels": new_labels}, res)
        return res

    def check_acceptance_criteria(self, fields: dict) -> StepResult:
        field_id = self.field_ids["acceptance_criteria"]
        body = fields.get(field_id) or ""
        keywords = self.cfg["acceptance_criteria_keywords"]
        missing = [kw for kw in keywords if not re.search(rf"\b{re.escape(kw)}\b", body, re.IGNORECASE)]
        expected_str = f"Body containing {keywords}"
        actual = body.strip() if body.strip() else "<empty>"
        if body.strip() and not missing:
            return StepResult("Acceptance Criteria", True, f"Contains {keywords}",
                              expected=expected_str, actual=actual)
        msg = "Empty" if not body.strip() else f"Missing parts: {missing}"
        return StepResult("Acceptance Criteria", False, msg,
                          expected=expected_str, actual=actual)

    def check_compliance_type(self, key: str, fields: dict) -> StepResult:
        return self._check_single_option_field(
            key, fields, "Compliance type",
            self.field_ids["compliance_type"],
            self.expected["compliance_type"],
        )

    def check_epic_link(self, fields: dict) -> StepResult:
        field_id = self.field_ids["epic_link"]
        v = fields.get(field_id)
        expected = "<non-empty Epic key>"
        actual = str(v) if _has_value(v) else "<empty>"
        if _has_value(v):
            return StepResult("Epic Link", True, f"Filled={v}",
                              expected=expected, actual=actual)
        return StepResult("Epic Link", False, "Empty or missing",
                          expected=expected, actual=actual)

    def check_team_name(self, key: str, fields: dict) -> StepResult:
        return self._check_single_option_field(
            key, fields, "Team Name",
            self.field_ids["team_name"],
            self.expected["team_name"],
        )

    def check_requirement_id(self, fields: dict) -> StepResult:
        field_id = self.field_ids["requirement_id"]
        v = fields.get(field_id)
        expected = "<non-empty Requirement ID>"
        actual = str(v) if _has_value(v) else "<empty>"
        if _has_value(v):
            return StepResult("Requirement ID", True, f"Filled={v}",
                              expected=expected, actual=actual)
        return StepResult("Requirement ID", False, "Empty or missing",
                          expected=expected, actual=actual)

    def check_region(self, key: str, fields: dict) -> StepResult:
        return self._check_single_option_field(
            key, fields, "Region",
            self.field_ids["region"],
            self.expected["region"],
        )

    def check_sector(self, key: str, fields: dict) -> StepResult:
        return self._check_single_option_field(
            key, fields, "Sector",
            self.field_ids["sector"],
            self.expected["sector"],
        )

    def check_description(self, fields: dict) -> StepResult:
        body = fields.get("description") or ""
        keywords = self.cfg["description_keywords"]
        missing = [kw for kw in keywords if not re.search(rf"\b{re.escape(kw)}\b", body, re.IGNORECASE)]
        expected_str = f"Body containing {keywords}"
        actual = body.strip() if body.strip() else "<empty>"
        if body.strip() and not missing:
            return StepResult("Description", True, f"Contains {keywords}",
                              expected=expected_str, actual=actual)
        msg = "Empty" if not body.strip() else f"Missing parts: {missing}"
        return StepResult("Description", False, msg,
                          expected=expected_str, actual=actual)

    def check_issue_links(self, fields: dict) -> StepResult:
        """'is child task of' target should match the Requirement ID / parent URL."""
        req_id_raw = fields.get(self.field_ids["requirement_id"])
        req_id = str(req_id_raw).strip() if req_id_raw is not None else ""

        parent_url = fields.get("customfield_10700") or ""
        parent_key_from_url = ""
        m = re.search(r"/browse/([A-Z]+-\d+)", parent_url)
        if m:
            parent_key_from_url = m.group(1)

        child_of_keys: list[str] = []
        for link in fields.get("issuelinks") or []:
            link_type = (link.get("type") or {}).get("inward", "")
            if link_type.lower() == "is child task of" and link.get("inwardIssue"):
                child_of_keys.append(link["inwardIssue"]["key"])

        expected_str = (
            f"'is child task of' target == RequirementID/Parent "
            f"({req_id or parent_key_from_url or '<unknown>'})"
        )
        actual = ", ".join(child_of_keys) if child_of_keys else "<no child-of link>"

        if not child_of_keys:
            return StepResult("Issue Links", False, "No 'is child task of' link found",
                              expected=expected_str, actual=actual)

        # Discrepancy check: child-of target must appear either in parent URL
        # (customfield_10700) or match the Requirement ID value.
        candidates = {parent_key_from_url, req_id}
        candidates.discard("")
        matched = any(ck in candidates for ck in child_of_keys) if candidates else False

        if matched:
            return StepResult("Issue Links", True,
                              f"Child-of {child_of_keys} matches requirement",
                              expected=expected_str, actual=actual)

        msg = (
            f"Child-of={child_of_keys}, RequirementID={req_id!r}, "
            f"ParentKeyFromURL={parent_key_from_url!r}"
        )
        return StepResult("Issue Links", False, f"Discrepancy: {msg}",
                          expected=expected_str, actual=actual)

    # --- orchestration ------------------------------------------------------

    def process(self, issue_key: str, team: str) -> IssueResult:
        started = time.perf_counter()
        result = IssueResult(issue_key=issue_key, team=team)

        try:
            issue = self.jira.get_issue(issue_key)
        except Exception as e:  # noqa: BLE001
            result.runtime_error = f"GET issue failed: {e}"
            result.overall = "FAIL"
            result.duration_seconds = round(time.perf_counter() - started, 3)
            return result

        if self.raw_response_dir is not None:
            try:
                out_path = self.raw_response_dir / f"{issue_key}.json"
                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(issue, f, indent=2, ensure_ascii=False)
            except Exception as e:  # noqa: BLE001
                log.warning("Failed to save raw response for %s: %s", issue_key, e)

        fields = issue.get("fields", {})

        steps = [
            self.check_affects_version(issue_key, fields),
            self.check_fix_version(issue_key, fields),
            self.check_component(issue_key, fields),
            self.check_labels(issue_key, fields, team),
            self.check_acceptance_criteria(fields),
            self.check_compliance_type(issue_key, fields),
            self.check_epic_link(fields),
            self.check_team_name(issue_key, fields),
            self.check_requirement_id(fields),
            self.check_region(issue_key, fields),
            self.check_sector(issue_key, fields),
            self.check_description(fields),
            self.check_issue_links(fields),
        ]
        result.steps = steps
        result.overall = "PASS" if all(s.passed for s in steps) else "FAIL"
        result.duration_seconds = round(time.perf_counter() - started, 3)
        return result


# ---------------------------------------------------------------------------
# Consolidated leader emails
# ---------------------------------------------------------------------------

def _recommendation_for(step: StepResult) -> str:
    """Short, human-friendly remediation text for a failing step."""
    if step.action.startswith("updated:"):
        return "Auto-fixed by the script. Please verify the change in Jira."
    if step.action.startswith("update-failed"):
        return "Auto-update failed; please apply the change manually."
    name = step.name
    if name == "Acceptance Criteria":
        return "Add the missing parts (Given / When / Then) to the Acceptance Criteria."
    if name == "Description":
        return "Add the missing parts (Who / What / Why) to the Description."
    if name == "Epic Link":
        return "Set the Epic Link in Jira (Details panel)."
    if name == "Requirement ID":
        return "Set the Requirement ID in Jira (Details panel)."
    if name == "Issue Links":
        return "Reconcile the 'is child task of' link with the Requirement ID."
    if name in ("Compliance type", "Team Name", "Region", "Sector"):
        return f"Verify current value and update to '{step.expected}' if appropriate."
    if name in ("Affects version", "Fix version", "Component", "Labels"):
        return "Add the expected value in Jira."
    return "Review the field manually in Jira."


def _html_escape(s: str) -> str:
    return (str(s)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


def _build_team_email(team: str, base_url: str,
                      team_results: list[IssueResult]) -> tuple[str, str, str]:
    """Build (subject, plain_text, html) for a team's consolidated summary."""
    rows: list[tuple[str, str, str, str, str]] = []  # key, step, expected, actual, action
    runtime_errors: list[tuple[str, str]] = []
    for r in team_results:
        if r.runtime_error and not r.steps:
            runtime_errors.append((r.issue_key, r.runtime_error))
            continue
        for s in r.steps:
            if not s.passed:
                rows.append((
                    r.issue_key, s.name, s.expected, s.actual, _recommendation_for(s),
                ))

    issue_count = len({k for k, *_ in rows} | {k for k, _ in runtime_errors})
    subject = (
        f"[Jira US Ready Check] {team}: {len(rows)} validation issue(s) "
        f"across {issue_count} user story/stories"
    )

    # Plain text
    text_lines = [
        f"Hi,",
        "",
        f"The Jira ready-check automation found {len(rows)} validation issue(s) "
        f"in {issue_count} user story/stories for the '{team}' team.",
        "",
        f"{'Issue':<14}{'Step':<22}{'Expected':<32}{'Actual':<32}Action",
        "-" * 130,
    ]
    for k, st, exp, act, action in rows:
        text_lines.append(f"{k:<14}{st:<22}{str(exp)[:30]:<32}{str(act)[:30]:<32}{action}")
    if runtime_errors:
        text_lines += ["", "Runtime errors:"]
        for k, err in runtime_errors:
            text_lines.append(f"  {k}: {err}")
    text = "\n".join(text_lines)

    # HTML
    def _link(key: str) -> str:
        return f'<a href="{base_url}/browse/{key}">{key}</a>'

    body_rows = "\n".join(
        f"<tr>"
        f"<td>{_link(_html_escape(k))}</td>"
        f"<td>{_html_escape(st)}</td>"
        f"<td>{_html_escape(exp)}</td>"
        f"<td>{_html_escape(act)}</td>"
        f"<td>{_html_escape(action)}</td>"
        f"</tr>"
        for k, st, exp, act, action in rows
    )
    runtime_block = ""
    if runtime_errors:
        rt_rows = "\n".join(
            f"<tr><td>{_link(_html_escape(k))}</td>"
            f"<td colspan='4'><i>Runtime error:</i> {_html_escape(err)}</td></tr>"
            for k, err in runtime_errors
        )
        runtime_block = (
            f"<h3 style='margin-top:24px;'>Runtime errors</h3>"
            f"<table border='1' cellpadding='6' cellspacing='0' "
            f"style='border-collapse:collapse;'>{rt_rows}</table>"
        )

    html = f"""\
<html>
  <body style="font-family: Arial, sans-serif; font-size: 13px; color:#222;">
    <p>Hi,</p>
    <p>
      The Jira ready-check automation found
      <b>{len(rows)}</b> validation issue(s) across
      <b>{issue_count}</b> user story/stories for the
      <b>{_html_escape(team)}</b> team. Please review and resolve the items below.
    </p>
    <table border="1" cellpadding="6" cellspacing="0"
           style="border-collapse:collapse; font-size:12px;">
      <thead style="background-color:#f0f0f0;">
        <tr>
          <th>Issue</th>
          <th>Failing step</th>
          <th>Expected value</th>
          <th>Actual value (data missing / wrong)</th>
          <th>Action to take</th>
        </tr>
      </thead>
      <tbody>
        {body_rows}
      </tbody>
    </table>
    {runtime_block}
    <p style="color:#888; font-size:11px;">
      Automatically generated by jira_checker.py.
    </p>
  </body>
</html>"""
    return subject, text, html


def send_consolidated_emails(mailer: Mailer, base_url: str,
                             teams_cfg: dict[str, Any],
                             results: list[IssueResult],
                             dry_run: bool) -> None:
    """Group failing/error results by team and send one summary email per team."""
    by_team: dict[str, list[IssueResult]] = {}
    for r in results:
        is_failure = bool(r.runtime_error) or any(not s.passed for s in r.steps)
        if not is_failure:
            continue
        by_team.setdefault(r.team, []).append(r)

    if not by_team:
        log.info("No failures to report; skipping team summary emails.")
        return

    for team, team_results in by_team.items():
        info = teams_cfg.get(team) or {}
        leader = info.get("leader_email")
        subject, text, html = _build_team_email(team, base_url, team_results)
        if not leader:
            log.warning("No leader_email configured for team '%s'; skipping summary email "
                        "(%d failing issue(s) not delivered).", team, len(team_results))
            _annotate_action(team_results, "no-leader-email-configured")
            continue
        if dry_run:
            log.info("[DRY-RUN team email] team=%s leader=%s issues=%d",
                     team, leader, len(team_results))
            _annotate_action(team_results, f"team-email-skipped(dry-run):{leader}")
            continue
        try:
            mailer.send(leader, subject, text, html_body=html)
            log.info("Sent summary email to %s for team '%s' (%d issue(s))",
                     leader, team, len(team_results))
            _annotate_action(team_results, f"team-email-sent:{leader}")
        except Exception as e:  # noqa: BLE001
            log.error("Failed to send summary email to %s: %s", leader, e)
            _annotate_action(team_results, f"team-email-failed:{e}")


def _annotate_action(team_results: list[IssueResult], suffix: str) -> None:
    """Append the team-email outcome to each failing step's action column."""
    for r in team_results:
        for s in r.steps:
            if not s.passed:
                s.action = f"{s.action}; {suffix}" if s.action else suffix


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

def read_input(path: Path) -> list[tuple[str, str]]:
    """Return a list of (issue_key, team). Team is the sheet name unless a
    'Team' column overrides it."""
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[tuple[str, str]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
        try:
            key_idx = next(i for i, h in enumerate(header) if h.lower() in ("issuekey", "issue key", "key"))
        except StopIteration:
            log.warning("Sheet %r has no IssueKey column; skipping", sheet_name)
            continue
        team_idx = next((i for i, h in enumerate(header) if h.lower() == "team"), None)

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[key_idx] is None:
                continue
            key = str(row[key_idx]).strip()
            if not key:
                continue
            team = sheet_name
            if team_idx is not None and row[team_idx]:
                team = str(row[team_idx]).strip()
            rows.append((key, team))
    return rows


def write_report(results: list[IssueResult], output_path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append([
        "Issue Key", "Team", "Overall", "Duration (s)", "Failing Steps",
        "Runtime Error",
    ])
    for r in results:
        summary.append([
            r.issue_key,
            r.team,
            r.overall,
            r.duration_seconds,
            "; ".join(r.failing_step_names()),
            r.runtime_error,
        ])

    details = wb.create_sheet("Details")
    details.append([
        "Issue Key", "Team", "Step", "Passed", "Expected Value", "Actual Value",
        "Message", "Action", "Error",
    ])
    for r in results:
        if r.runtime_error and not r.steps:
            details.append([r.issue_key, r.team, "-", "FAIL", "-", "-", "-", "-", r.runtime_error])
            continue
        for s in r.steps:
            details.append([
                r.issue_key, r.team, s.name, "PASS" if s.passed else "FAIL",
                s.expected, s.actual, s.message, s.action, s.error,
            ])

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Jira US ready-check automation")
    parser.add_argument("--input", "-i", required=True, help="Input Excel file with user stories")
    parser.add_argument("--config", "-c", default="config.json", help="Path to config.json")
    parser.add_argument("--output", "-o", default=None, help="Output report path (default: report_<ts>.xlsx)")
    parser.add_argument("--no-update", action="store_true", help="Do not push field updates to Jira")
    parser.add_argument("--no-email", action="store_true", help="Do not send emails to leaders")
    parser.add_argument("--dry-run", action="store_true", help="Equivalent to --no-update --no-email")
    args = parser.parse_args(argv)

    input_path = Path(args.input)
    config_path = Path(args.config)
    if not input_path.exists():
        log.error("Input file not found: %s", input_path)
        return 2
    if not config_path.exists():
        log.error("Config file not found: %s", config_path)
        return 2

    config = load_config(config_path)
    token = os.getenv("JIRA_BEARER_TOKEN")
    if not token:
        log.error("JIRA_BEARER_TOKEN not set (env or .env file)")
        return 2

    auto_update = not (args.no_update or args.dry_run)
    send_emails = not (args.no_email or args.dry_run)

    jira = JiraClient(
        base_url=config["jira"]["base_url"],
        bearer_token=token,
        verify_ssl=config["jira"].get("verify_ssl", True),
        timeout=config["jira"].get("timeout_seconds", 30),
    )
    mailer = Mailer(config["smtp"], dry_run=not send_emails)

    run_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)
    raw_dir = reports_dir / f"raw_{run_ts}"
    processor = IssueProcessor(
        jira, mailer, config,
        auto_update=auto_update, send_emails=send_emails,
        raw_response_dir=raw_dir,
    )

    stories = read_input(input_path)
    log.info("Processing %d user stories", len(stories))

    results: list[IssueResult] = []
    for idx, (key, team) in enumerate(stories, start=1):
        log.info("[%d/%d] %s (team=%s)", idx, len(stories), key, team)
        try:
            res = processor.process(key, team)
        except Exception:  # noqa: BLE001 - never let one issue stop the batch
            res = IssueResult(
                issue_key=key, team=team, overall="FAIL",
                runtime_error=traceback.format_exc(limit=3),
            )
        log.info("  -> %s in %.2fs (failing=%s)",
                 res.overall, res.duration_seconds, res.failing_step_names())
        results.append(res)

    send_consolidated_emails(
        mailer=mailer,
        base_url=config["jira"]["base_url"],
        teams_cfg=config["teams"],
        results=results,
        dry_run=not send_emails,
    )

    if args.output:
        out = Path(args.output)
    else:
        out = reports_dir / f"report_{run_ts}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    write_report(results, out)
    log.info("Report written: %s", out)
    if raw_dir.exists():
        log.info("Raw API responses: %s", raw_dir)

    fails = sum(1 for r in results if r.overall != "PASS")
    log.info("Done. %d PASS, %d FAIL", len(results) - fails, fails)
    return 0 if fails == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
