"""Generates a sample input Excel file with the structure the checker expects."""
from openpyxl import Workbook

wb = Workbook()

ws_fin = wb.active
ws_fin.title = "Finance"
ws_fin.append(["IssueKey", "Team"])
ws_fin.append(["AASQ-72454", "Finance"])

ws_del = wb.create_sheet("Delivery")
ws_del.append(["IssueKey", "Team"])

ws_loa = wb.create_sheet("Loaner")
ws_loa.append(["IssueKey", "Team"])

wb.save("input_user_stories.xlsx")
print("Created input_user_stories.xlsx")
